import sys, os, openpyxl
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QHeaderView, QPushButton, QLabel, QSplitter, QTableWidget, 
    QTableWidgetItem, QMessageBox, QFileDialog, QLineEdit, QComboBox, 
    QFrame, QCheckBox
)
from PySide6.QtCore import Qt, QFile, QTextStream
from PySide6.QtGui import QColor, QPainter
from openpyxl.utils import coordinate_to_tuple

# IMPORT DARI STRUKTUR FOLDER ANDA
from viewmodels.mapper_vm import MapperViewModel
from models.excel_handler import ExcelTableModel
from views.components.excel_grid import DraggableTableView, DroppableTableView

class LoadingOverlay(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WA_TransparentForMouseEvents)
        self.hide()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.fillRect(self.rect(), QColor(0, 0, 0, 150)) 
        painter.setPen(Qt.white)
        font = painter.font()
        font.setPointSize(22)
        font.setBold(True)
        painter.setFont(font)
        painter.drawText(self.rect(), Qt.AlignCenter, "⏳ Membaca Excel...\nMohon Tunggu")

class MainWindow(QMainWindow):
    def __init__(self, view_model: MapperViewModel):
        super().__init__()
        self.vm = view_model
        
        self.src_matches = []
        self.dest_matches = []
        self.src_search_idx = -1
        self.dest_search_idx = -1

        self.setWindowTitle("Excel Mapper Pro - Project Magang")
        self.resize(1350, 850)
        
        # Load CSS Eksternal
        self.load_stylesheet("views/style.css")
        
        self.setup_ui()
        self.setup_connections()
        self.overlay = LoadingOverlay(self)

    def load_stylesheet(self, filename=None):
        import os
        base_dir = os.path.dirname(os.path.abspath(__file__))
        full_path = os.path.join(base_dir, "web", "css", "style.css")
        
        file = QFile(full_path)
        if file.open(QFile.ReadOnly | QFile.Text):
            self.setStyleSheet(QTextStream(file).readAll())
            print(f"✅ Style loaded: {full_path}")
        else:
            print(f"❌ Style NOT found: {full_path}")

    def setup_ui(self):
        central_widget = QWidget(objectName="central")
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # --- HEADER AREA ---
        header_layout = QHBoxLayout()
        title_label = QLabel("🔗 Excel Mapper Tool")
        title_label.setStyleSheet("font-size: 22px; font-weight: bold; color: #1e3a8a;")
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        self.btn_undo = QPushButton("↩ Undo"); self.btn_undo.setProperty("class", "icon_btn")
        self.btn_load_preset = QPushButton("📂 Load Preset"); self.btn_load_preset.setProperty("class", "icon_btn")
        self.btn_save_preset = QPushButton("💾 Save Preset"); self.btn_save_preset.setProperty("class", "icon_btn")
        self.chk_auto_append = QCheckBox("Auto-Append")
        self.chk_auto_append.setStyleSheet("font-weight: bold; margin-right: 10px;")
        self.btn_run = QPushButton("▶ Run Mapping"); self.btn_run.setObjectName("btn_run")

        for w in [self.btn_undo, self.btn_load_preset, self.btn_save_preset, self.chk_auto_append, self.btn_run]:
            header_layout.addWidget(w)
        main_layout.addLayout(header_layout)

        # --- CONFIG PANEL ---
        config_frame = QFrame(objectName="card")
        config_layout = QHBoxLayout(config_frame)
        for p_type in ["src", "dest"]:
            box = QVBoxLayout()
            box.addWidget(QLabel("SOURCE DATA" if p_type == "src" else "DESTINATION DATA", 
                                styleSheet="font-weight: bold; color: #64748b; font-size: 11px;"))
            row = QHBoxLayout()
            f_in = QLineEdit(readOnly=True, placeholderText="Pilih file...")
            b_br = QPushButton("Browse"); b_br.setProperty("class", "action_style")
            s_cm = QComboBox()
            row.addWidget(f_in, 3); row.addWidget(b_br); row.addWidget(s_cm, 1)
            box.addLayout(row)
            if p_type == "src":
                self.src_file_input, self.btn_src_browse, self.src_sheet_combo = f_in, b_br, s_cm
                config_layout.addLayout(box, 1)
                config_layout.addWidget(QLabel("➔", styleSheet="font-size: 24px; color: #cbd5e1;"))
            else:
                self.dest_file_input, self.btn_dest_browse, self.dest_sheet_combo = f_in, b_br, s_cm
                config_layout.addLayout(box, 1)
        main_layout.addWidget(config_frame)

        # --- WORKSPACE DENGAN SPLITTER ---
        self.main_splitter = QSplitter(Qt.Vertical) # Splitter Utama (Atas-Bawah)
        self.upper_splitter = QSplitter(Qt.Horizontal) # Splitter Workspace (Kiri-Kanan)

        for p_type in ["src", "dest"]:
            frame = QFrame(objectName="card")
            vbox = QVBoxLayout(frame)
            
            h_header = QHBoxLayout()
            h_header.addWidget(QLabel("🔵 Source Sheet" if p_type == "src" else "🟢 Destination Sheet", 
                                    styleSheet="font-weight: bold; color: #1e3a8a;"))
            h_header.addStretch()
            tg, zo, zi = QPushButton("👁️"), QPushButton("-"), QPushButton("+")
            for b in [tg, zo, zi]: 
                b.setProperty("class", "action_style")
                b.setFixedSize(35, 30)
                h_header.addWidget(b)
            
            s_lay = QHBoxLayout()
            s_in = QLineEdit(placeholderText="🔍 Cari data...")
            b_pr, b_nx = QPushButton("⇽"), QPushButton("⇾")
            for b in [b_pr, b_nx]: 
                b.setProperty("class", "action_style")
                b.setFixedSize(30, 30)
                s_lay.addWidget(b)
            s_lay.insertWidget(0, s_in)
            
            vbox.addLayout(h_header)
            vbox.addLayout(s_lay)
            
            if p_type == "src":
                self.btn_src_toggle_view, self.btn_src_zoom_out, self.btn_src_zoom_in = tg, zo, zi
                self.src_search_input, self.btn_src_prev, self.btn_src_next = s_in, b_pr, b_nx
                self.source_view = DraggableTableView()
                # PERBAIKAN UKURAN DEFAULT TABEL
                self.source_view.horizontalHeader().setDefaultSectionSize(120)
                self.source_view.verticalHeader().setMinimumSectionSize(55)
                vbox.addWidget(self.source_view)
                self.upper_splitter.addWidget(frame)
            else:
                self.btn_dest_toggle_view, self.btn_dest_zoom_out, self.btn_dest_zoom_in = tg, zo, zi
                self.dest_search_input, self.btn_dest_prev, self.btn_dest_next = s_in, b_pr, b_nx
                self.dest_view = DroppableTableView(self.handle_drop)
                # PERBAIKAN UKURAN DEFAULT TABEL
                self.dest_view.horizontalHeader().setDefaultSectionSize(120)
                self.dest_view.verticalHeader().setMinimumSectionSize(55)
                vbox.addWidget(self.dest_view)
                self.upper_splitter.addWidget(frame)

        # --- RULES TABLE AREA ---
        self.rules_frame = QFrame(objectName="card")
        r_layout = QVBoxLayout(self.rules_frame)
        r_layout.addWidget(QLabel("MAPPING RULES", styleSheet="font-weight: bold; color: #64748b; font-size: 11px;"))
        
        self.rules_table = QTableWidget(0, 4)
        self.rules_table.setObjectName("rules_table")
        self.rules_table.setHorizontalHeaderLabels(["Source", "Destination", "Type", "Action"])
        self.rules_table.verticalHeader().setVisible(False)
        self.rules_table.setShowGrid(False)
        self.rules_table.setAlternatingRowColors(True)
        self.rules_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # PERBAIKAN RATA TENGAH JUDUL
        self.rules_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.rules_table.horizontalHeader().setMinimumHeight(40)
        r_layout.addWidget(self.rules_table)

        # --- MERAKIT SPLITTER ---
        self.main_splitter.addWidget(self.upper_splitter) # Workspace masuk bagian atas
        self.main_splitter.addWidget(self.rules_frame)     # Rules masuk bagian bawah
        self.main_splitter.setStretchFactor(0, 3)          # Workspace lebih luas
        self.main_splitter.setStretchFactor(1, 1)

        main_layout.addWidget(self.main_splitter)

    def setup_connections(self):
        # File Browsing
        self.btn_src_browse.clicked.connect(lambda: self.browse_file(self.src_file_input, self.src_sheet_combo))
        self.btn_dest_browse.clicked.connect(lambda: self.browse_file(self.dest_file_input, self.dest_sheet_combo))
        self.src_sheet_combo.currentTextChanged.connect(lambda s: self.load_sheet(self.src_file_input.text(), s, self.source_view))
        self.dest_sheet_combo.currentTextChanged.connect(lambda s: self.load_sheet(self.dest_file_input.text(), s, self.dest_view))
        
        # Search & Navigation
        self.src_search_input.textChanged.connect(lambda: self.apply_highlight(self.source_view, "src"))
        self.dest_search_input.textChanged.connect(lambda: self.apply_highlight(self.dest_view, "dest"))
        self.btn_src_next.clicked.connect(lambda: self.navigate_search("src", True))
        self.btn_src_prev.clicked.connect(lambda: self.navigate_search("src", False))
        self.btn_dest_next.clicked.connect(lambda: self.navigate_search("dest", True))
        self.btn_dest_prev.clicked.connect(lambda: self.navigate_search("dest", False))

        # Main Actions
        self.btn_run.clicked.connect(self.run_mapping_with_current_files)
        self.btn_undo.clicked.connect(self.vm.undo_last_rule)
        self.btn_save_preset.clicked.connect(self.save_preset_dialog)
        self.btn_load_preset.clicked.connect(self.load_preset_dialog)

        # ViewModel Signals
        self.vm.rules_updated.connect(self.refresh_rules_ui)
        self.vm.mapping_started.connect(lambda: self.btn_run.setEnabled(False))
        self.vm.mapping_finished.connect(self.on_mapping_success)
        self.vm.mapping_error.connect(lambda msg: QMessageBox.critical(self, "Error", msg))

       # --- BAGIAN ZOOM & VIEW TOGGLE (VERSI RAPI) ---

        # Hapus semua baris zoom lama, ganti dengan 6 baris ini saja:
        self.btn_src_zoom_in.clicked.connect(lambda: self.source_view.apply_zoom(1))
        self.btn_src_zoom_out.clicked.connect(lambda: self.source_view.apply_zoom(-1))

        self.btn_dest_zoom_in.clicked.connect(lambda: self.dest_view.apply_zoom(1))
        self.btn_dest_zoom_out.clicked.connect(lambda: self.dest_view.apply_zoom(-1))
        self.vm.mapping_error.connect(self.on_mapping_error)
        
    def on_mapping_error(self, msg):
        QMessageBox.critical(self, "Error", msg)
        self.btn_run.setEnabled(True) # AKTIFKAN KEMBALI TOMBOLNYA
        
    def run_mapping_with_current_files(self):
        current_src = self.src_file_input.text()
        current_dest = self.dest_file_input.text()

        if not current_src or not current_dest:
            QMessageBox.warning(self, "Peringatan", "Pilih file Source dan Destination dulu!")
            return

        # Paksa semua rules menggunakan file yang baru Anda pilih di layar
        for rule in self.vm.rules:
            rule["src_file"] = current_src
            rule["dest_file"] = current_dest

        # Jalankan eksekusi
        self.vm.run_mapping(self.chk_auto_append.isChecked())

    def browse_file(self, line_edit, combo_box):
        path, _ = QFileDialog.getOpenFileName(self, "Buka Excel", "", "Excel (*.xlsx *.xlsm)")
        if path:
            line_edit.setText(os.path.normpath(path))
            wb = openpyxl.load_workbook(path, read_only=True)
            combo_box.clear(); combo_box.addItems(wb.sheetnames); wb.close()

    def load_sheet(self, filepath, sheetname, view):
        if filepath and sheetname:
            self.overlay.show(); QApplication.processEvents()
            try:
                # 1. Muat Model
                model = ExcelTableModel(filepath, sheetname)
                view.setModel(model)
                view.sync_with_excel()
                
                # 2. PAKSA UKURAN KOLOM (Kunci Perbaikan)
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb[sheetname]
                
                for i in range(1, ws.max_column + 1):
                    col_letter = openpyxl.utils.get_column_letter(i)
                    # Ambil lebar dari file asli
                    excel_width = ws.column_dimensions[col_letter].width
                    
                    # Logika: Jika lebar di excel tidak didefinisikan (None), gunakan 150px
                    # Jika ada, konversi ke pixel (excel_width * 10)
                    if excel_width is None or excel_width == 0:
                        final_width = 150 
                    else:
                        final_width = int(excel_width * 5)
                    
                    view.setColumnWidth(i-1, final_width)
                
                # 3. PAKSA TINGGI BARIS agar tidak sesak
                view.verticalHeader().setDefaultSectionSize(10)
                # Pastikan angka kolom kiri (1, 2, 3) punya ruang 60px
                view.verticalHeader().setMinimumSectionSize(50) 
                
                wb.close()
                self.refresh_highlights()
                
            except Exception as e:
                print(f"Error loading sheet: {e}")
            finally: 
                self.overlay.hide()

    def handle_drop(self, src_cell, dest_cell):
        src_model = self.source_view.model()
        r, c = coordinate_to_tuple(src_cell)
        val = src_model.grid_data[r-1][c-1]["value"] if src_model else "N/A"
        self.vm.add_rule(self.src_file_input.text(), self.src_sheet_combo.currentText(), src_cell, val,
                         self.dest_file_input.text(), self.dest_sheet_combo.currentText(), dest_cell)

    def refresh_rules_ui(self, rules):
        self.rules_table.setRowCount(0)
        for i, rule in enumerate(rules):
            row = self.rules_table.rowCount(); self.rules_table.insertRow(row)
            items = [QTableWidgetItem(f"📄 {rule['src_sheet']}!{rule['src_cell']}"), 
                     QTableWidgetItem(f"🎯 {rule['dest_sheet']}!{rule['dest_cell']}"), 
                     QTableWidgetItem("⚡ Live Write")]
            for col, item in enumerate(items):
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable); self.rules_table.setItem(row, col, item)
            btn_del = QPushButton("✕"); btn_del.setFixedSize(26, 26)
            btn_del.setObjectName("btn_delete_rule") # Menggunakan ID dari CSS
            btn_del.clicked.connect(lambda _, idx=i: self.vm.remove_rule(idx))
            c = QWidget(); l = QHBoxLayout(c); l.setContentsMargins(0,0,0,0); l.setAlignment(Qt.AlignCenter); l.addWidget(btn_del)
            self.rules_table.setCellWidget(row, 3, c)
        self.refresh_highlights()

    def apply_highlight(self, view, p_type):
        model = view.model()
        if not model: return
        text = self.src_search_input.text().lower() if p_type == "src" else self.dest_search_input.text().lower()
        matches = []
        model.layoutAboutToBeChanged.emit()
        for r in range(model.rowCount()):
            for c in range(model.columnCount()):
                idx = model.index(r, c)
                if text and text in str(model.grid_data[r][c]["value"]).lower():
                    model.setData(idx, QColor("#ccfbf1"), Qt.BackgroundRole); matches.append(idx)
                else: model.setData(idx, None, Qt.BackgroundRole)
        if p_type == "src": self.src_matches, self.src_search_idx = matches, -1
        else: self.dest_matches, self.dest_search_idx = matches, -1
        model.layoutChanged.emit()
        if matches: self.navigate_search(p_type, True)

    def navigate_search(self, p_type, forward):
        matches = self.src_matches if p_type == "src" else self.dest_matches
        view = self.source_view if p_type == "src" else self.dest_view
        if not matches: return
        if p_type == "src": self.src_search_idx = (self.src_search_idx + 1) % len(matches) if forward else (self.src_search_idx - 1) % len(matches); idx = matches[self.src_search_idx]
        else: self.dest_search_idx = (self.dest_search_idx + 1) % len(matches) if forward else (self.dest_search_idx - 1) % len(matches); idx = matches[self.dest_search_idx]
        view.scrollTo(idx); view.setCurrentIndex(idx)

    def refresh_highlights(self):
        src_model, dest_model = self.source_view.model(), self.dest_view.model()
        if src_model: src_model.clear_highlights()
        if dest_model: dest_model.clear_highlights()
        for rule in self.vm.rules:
            if src_model and rule["src_sheet"] == self.src_sheet_combo.currentText(): src_model.add_highlight(rule["src_cell"], QColor("#dbeafe"))
            if dest_model and rule["dest_sheet"] == self.dest_sheet_combo.currentText(): dest_model.add_highlight(rule["dest_cell"], QColor("#dcfce3"))

    def save_preset_dialog(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Preset", "", "JSON (*.json)")
        if path: self.vm.save_preset(path); QMessageBox.information(self, "Sukses", "Aturan disimpan!")

    def load_preset_dialog(self):
        path, _ = QFileDialog.getOpenFileName(self, "Load Preset", "", "JSON (*.json)")
        if path: self.vm.load_preset(path); QMessageBox.information(self, "Sukses", "Aturan dimuat!")

    def on_mapping_success(self, msg, count):
        self.btn_run.setEnabled(True)
        QMessageBox.information(self, "Berhasil", msg)

    def resizeEvent(self, event):
        if hasattr(self, 'overlay'): self.overlay.resize(self.size())
        super().resizeEvent(event)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    vm = MapperViewModel()
    window = MainWindow(vm)
    window.show()
    sys.exit(app.exec())