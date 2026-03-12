import openpyxl
from openpyxl.utils import column_index_from_string
from PySide6.QtCore import Qt, QAbstractTableModel
from PySide6.QtGui import QColor, QFont, QBrush

def get_excel_col_name(n):
    """Mengonversi indeks angka ke nama kolom Excel (0 -> A, 1 -> B)."""
    result = ""
    n += 1
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

class ExcelTableModel(QAbstractTableModel):
    def __init__(self, filepath, sheet_name):
        super().__init__()
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.grid_data = []
        self.max_row = 0
        self.max_col = 0
        self.merged_ranges = []
        self.col_widths = {}
        self.row_heights = {}
        self.hidden_cols = []  # Menyimpan indeks kolom yang disembunyikan di Excel asli
        self.highlighted_cells = {} 
        self.font_size = 10
        self._load_excel()

    def _load_excel(self):
        # Memuat workbook dengan data_only=True agar formula terbaca sebagai nilai.
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb[self.sheet_name]
        
        self.max_col = ws.max_column
        self.max_row = ws.max_row
        self.merged_ranges = ws.merged_cells.ranges
        
        # Deteksi kolom yang disembunyikan (Hidden Columns)
        self.hidden_cols = []
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.hidden:
                idx = column_index_from_string(col_letter) - 1
                self.hidden_cols.append(idx)
        
        self.grid_data = [[None for _ in range(self.max_col)] for _ in range(self.max_row)]
        
        for row in ws.iter_rows(min_row=1, max_row=self.max_row, min_col=1, max_col=self.max_col):
            for cell in row:
                r_idx, c_idx = cell.row - 1, cell.column - 1
                val = "" if cell.value is None else str(cell.value)
                bg, bold, align = None, False, Qt.AlignLeft | Qt.AlignVCenter
                
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
                    rgb = cell.fill.fgColor.rgb
                    if isinstance(rgb, str) and rgb != '00000000':
                        bg = f"#{rgb[2:]}" if len(rgb) == 8 else f"#{rgb}"
                
                if cell.font: bold = cell.font.bold
                
                if cell.alignment:
                    if cell.alignment.horizontal == 'center': align = Qt.AlignCenter
                    elif cell.alignment.horizontal == 'right': align = Qt.AlignRight | Qt.AlignVCenter

                self.grid_data[r_idx][c_idx] = {
                    "value": val, "bg": bg, "bold": bold, "align": align
                }
                
        for col_letter, col_dim in ws.column_dimensions.items():
            idx = column_index_from_string(col_letter) - 1
            if col_dim.width: self.col_widths[idx] = col_dim.width
                
        for r_idx, r_dim in ws.row_dimensions.items():
            if r_dim.height: self.row_heights[r_idx - 1] = r_dim.height
        wb.close()

    def rowCount(self, parent=None): return self.max_row
    def columnCount(self, parent=None): return self.max_col

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid(): return None
        r, c = index.row(), index.column()
        cell_data = self.grid_data[r][c]
        
        if role == Qt.DisplayRole: return cell_data["value"]
        if role == Qt.TextAlignmentRole: return cell_data["align"]
        if role == Qt.BackgroundRole:
            cell_ref = f"{get_excel_col_name(c)}{r + 1}"
            if cell_ref in self.highlighted_cells: return QBrush(self.highlighted_cells[cell_ref])
            if cell_data["bg"]: return QBrush(QColor(cell_data["bg"]))
        if role == Qt.FontRole:
            font = QFont()
            font.setPointSize(self.font_size)
            font.setBold(cell_data["bold"])
            return font
        if role == Qt.ForegroundRole: return QBrush(QColor("black")) 
        return None

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole: return get_excel_col_name(col)
        if orientation == Qt.Vertical and role == Qt.DisplayRole: return str(col + 1)
        return None

    def set_font_size(self, size):
        self.font_size = size
        self.layoutChanged.emit()

    def add_highlight(self, cell_ref, color):
        self.highlighted_cells[cell_ref] = color
        self.layoutChanged.emit()

    def clear_highlights(self):
        self.highlighted_cells.clear()
        self.layoutChanged.emit()